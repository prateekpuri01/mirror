"""Import jobs from the user's Job Hunt 2026.xlsx spreadsheet as positive interest signals.

Usage (inside Docker):
    python -m app.scripts.import_spreadsheet

Or from host:
    docker compose exec api python -m app.scripts.import_spreadsheet
"""

import asyncio
import logging
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.database import async_session
from app.models.jobs import Job, JobSource

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

SPREADSHEET_PATH = "/app/docs/Job Hunt 2026.xlsx"


def parse_spreadsheet(path: str) -> list[dict]:
    """Parse the xlsx file and return a list of job dicts."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # First row is header
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    jobs = []

    for row in rows[1:]:
        values = list(row)
        if not values or not values[0]:  # Skip empty rows
            continue

        entry = {}
        for i, col_name in enumerate(header):
            if i < len(values):
                entry[col_name] = values[i]

        company = str(entry.get("company", entry.get("company ", ""))).strip()
        title = str(entry.get("job title", "")).strip()

        if not company or not title:
            continue

        job_data = {
            "company": company,
            "title": title,
            "location": str(entry.get("location", "")).strip() or None,
            "url": str(entry.get("job posting link", "")).strip() or None,
            "salary_text": str(entry.get("salary?", "")).strip() or None,
        }
        jobs.append(job_data)

    wb.close()
    return jobs


def parse_salary(salary_text: str | None) -> tuple[int | None, int | None]:
    """Extract salary min/max from text like '$350,000 - $500,000'."""
    if not salary_text:
        return None, None

    import re

    amounts = re.findall(r"\$?([\d,]+)", salary_text)
    if not amounts:
        return None, None

    values = [int(a.replace(",", "")) for a in amounts]
    # Fix truncated values like $350,00 → $350,000
    values = [v * 10 if 10000 < v < 100000 else v for v in values]

    if len(values) >= 2:
        return min(values), max(values)
    return values[0], values[0]


async def import_jobs():
    """Import spreadsheet jobs, matching existing or creating new entries."""
    path = Path(SPREADSHEET_PATH)
    if not path.exists():
        logger.error("Spreadsheet not found at %s", SPREADSHEET_PATH)
        return

    entries = parse_spreadsheet(SPREADSHEET_PATH)
    logger.info("Parsed %d jobs from spreadsheet", len(entries))

    async with async_session() as session:
        matched = 0
        created = 0

        for entry in entries:
            company = entry["company"]
            title = entry["title"]

            # Try fuzzy match on existing jobs
            pattern_company = f"%{company}%"
            pattern_title = f"%{title.split()[0]}%"  # Match first word of title

            result = await session.execute(
                select(Job)
                .where(
                    Job.company.ilike(pattern_company),
                    Job.title.ilike(pattern_title),
                )
                .limit(1)
            )
            existing = result.scalar_one_or_none()

            if existing:
                existing.thumbs = 1
                matched += 1
                logger.info("Matched: %s at %s → thumbs=1", title, company)
            else:
                # Create new job entry
                salary_min, salary_max = parse_salary(entry.get("salary_text"))
                url = (
                    entry.get("url")
                    or f"manual://{company.lower().replace(' ', '-')}/{title.lower().replace(' ', '-')}"
                )

                # Check if URL already exists
                url_check = await session.execute(select(Job).where(Job.url == url).limit(1))
                existing_by_url = url_check.scalar_one_or_none()
                if existing_by_url:
                    existing_by_url.thumbs = 1
                    matched += 1
                    continue

                new_job = Job(
                    title=title,
                    company=company,
                    location=entry.get("location"),
                    remote=entry.get("location", "").lower() == "remote"
                    if entry.get("location")
                    else False,
                    salary_min=salary_min,
                    salary_max=salary_max,
                    description=f"Manually tracked job: {title} at {company}",
                    url=url,
                    source=JobSource.manual,
                    thumbs=1,
                )
                session.add(new_job)
                created += 1
                logger.info("Created: %s at %s (source=manual, thumbs=1)", title, company)

        await session.commit()
        logger.info("Import complete: %d matched existing, %d created new", matched, created)


if __name__ == "__main__":
    asyncio.run(import_jobs())
